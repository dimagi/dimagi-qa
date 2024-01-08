import os
import time
import pandas as pd
from datetime import datetime, timedelta

import requests
from openpyxl import load_workbook
from selenium.webdriver import ActionChains

from common_utilities.selenium.base_page import BasePage
from common_utilities.path_settings import PathSettings
from HQSmokeTests.userInputs.user_inputs import UserData
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, \
    ElementClickInterceptedException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

""""Contains test page elements and functions related to the exports module"""

#
# def latest_download_file():
#     os.chdir(PathSettings.DOWNLOAD_PATH)
#     files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
#     newest = max(files, key=os.path.getctime)
#     print("File downloaded: " + newest)
#     return newest

def latest_download_file(type=".xlsx"):
    cwd = os.getcwd()
    try:
        os.chdir(PathSettings.DOWNLOAD_PATH)
        all_specific_files = filter(lambda x: x.endswith(type), os.listdir(os.getcwd()))
        files = sorted(all_specific_files, key=os.path.getctime)
        if files[-1].endswith(".log"):
            newest = sorted(files, key=os.path.getctime)[-2]
        elif files[-1].endswith(".xlsx"):
            newest = sorted(files, key=os.path.getctime)[-1]
        else:
            newest = max(files, key=os.path.getctime)
        print("File downloaded: " + newest)
        return newest
    finally:
        print("Restoring the path...")
        os.chdir(cwd)
        print("Current directory is-", os.getcwd())


class ExportDataPage(BasePage):

    def __init__(self, driver):
        super().__init__(driver)

        self.presentday = datetime.now()  # or presentday = datetime.today()
        self.date_having_submissions = "2022-01-18" + " to " + datetime.now().strftime('%Y-%m-%d')
        # Get Yesterday
        self.yesterday = self.presentday - timedelta(1)
        self.nextday = self.presentday + timedelta(1)
        self.current_date_range = self.yesterday.strftime('%Y-%m-%d') + " to " + datetime.now().strftime('%Y-%m-%d')
        self.next_date_range = self.yesterday.strftime('%Y-%m-%d') + " to " + self.nextday.strftime('%Y-%m-%d')

        # Add Export
        self.data_dropdown = (By.LINK_TEXT, 'Data')
        self.view_all_link = (By.LINK_TEXT, 'View All')
        self.add_export_button = (By.XPATH, "//a[@href='#createExportOptionsModal']")
        self.add_export_conf = (By.XPATH, "//button[@data-bind='visible: showSubmit, disable: disableSubmit']")
        self.export_name = (By.XPATH, '//*[@id="export-name"]')
        self.export_settings_create = (By.XPATH, "//button[@class='btn btn-lg btn-primary']")
        self.date_range = (By.ID, "id_date_range")
        self.case_owner = (By.XPATH, "//span[@class='select2-selection select2-selection--multiple']")

        # Export Form and Case data variables
        self.export_form_data_link = (By.LINK_TEXT, 'Export Form Data')
        self.export_case_data_link = (By.LINK_TEXT, 'Export Case Data')
        self.export_form_case_data_button = (By.XPATH, "(//a[@class='btn btn-primary'])[2]")
        self.web_users_option = (By.XPATH, "//li/span[.='[Web Users]']")
        self.all_data_option = (By.XPATH, "//li/span[.='[All Data]']")
        self.users_field = (By.XPATH, "(//textarea[@class='select2-search__field'])[1]")
        self.user_from_list = "//li[contains(.,'{}')]"
        self.prepare_export_button = (By.XPATH, "//button[@data-bind='disable: disablePrepareExport']")
        self.download_button = (By.XPATH, "//a[@class='btn btn-primary btn-full-width']")
        self.apply = (By.XPATH, "//button[@class='applyBtn btn btn-sm btn-primary']")
        self.export_button = (By.XPATH, "//a[@class='btn btn-primary'][contains(text(),'Export')]")

        # Find Data By ID
        self.find_data_by_ID = (By.LINK_TEXT, 'Find Data by ID')
        self.find_data_by_ID_textbox = (By.XPATH, "//input[@placeholder='Form Submission ID']")
        self.find_data_by_case_ID_textbox = (By.XPATH, "//input[@placeholder='Case ID']")
        self.find_data_by_case_ID_button = (By.XPATH, "//div[./input[@placeholder='Case ID']]//following-sibling::div/button")
        self.find_data_by_ID_button = (By.XPATH, "//div[./input[@placeholder='Form Submission ID']]//following-sibling::div/button")
        self.view_FormID_CaseID = (By.LINK_TEXT, 'View')
        self.case_id_value = "//th[contains(.,'Case ID')]//following-sibling::td[contains(.,'{}')]"
        self.related_cases_tab = (By.LINK_TEXT, "Related Cases")
        self.related_cases_view = (By.XPATH, "//td[contains(.,'"+UserData.child_name+"')]//following-sibling::div/a[contains(.,'View')]")
        self.woman_form_name_HQ = (By.XPATH, "(//div[@class='form-data-readable form-data-raw'])[1]")
        self.woman_case_name_HQ = (By.XPATH, "//th[@title='name']//following::td[1]")

        # Export SMS variables
        self.export_sms_link = (By.LINK_TEXT, "Export SMS Messages")

        # Daily Saved Export variables, form, case
        self.daily_saved_export_link = (By.LINK_TEXT, 'Daily Saved Exports')
        self.edit_form_case_export = (By.XPATH, "(//a[contains(@data-bind,'edit')])[1]")
        self.create_DSE_checkbox = (By.XPATH, '//*[@id="daily-saved-export-checkbox"]')
        self.download_dse = (By.XPATH, "(//a[@class='btn btn-info btn-xs'])[1]")
        self.download_dse_form = (By.XPATH,
                                  "//h4[.//span[.='" + UserData.form_export_name_dse + "']]/following-sibling::div//*[./i[@class='fa fa-cloud-download']]")
        self.download_dse_case = (By.XPATH,
                                  "//h4[.//span[.='" + UserData.case_export_name_dse + "']]/following-sibling::div//*[./i[@class='fa fa-cloud-download']]")

        self.data_upload_msg = (By.XPATH, "//*[contains(text(),'Data update complete')]")
        self.data_upload_msg_form = (By.XPATH,
                                     "//h4[.//span[.='" + UserData.form_export_name_dse + "']]/following-sibling::div//*[contains(text(),'Data update complete')]")
        self.data_upload_msg_case = (By.XPATH,
                                     "//h4[.//span[.='" + UserData.case_export_name_dse + "']]/following-sibling::div//*[contains(text(),'Data update complete')]")

        # Excel Dashboard Integrations, form, case
        self.export_excel_dash_int = (By.LINK_TEXT, 'Excel Dashboard Integration')
        self.update_data = (By.XPATH, "//button[@data-toggle='modal'][1]")
        self.update_data_conf = (By.XPATH, "//button[@data-bind='click: emailedExport.updateData']")

        self.update_data_form = (By.XPATH,
                                 "//h4[.//span[.='" + UserData.form_export_name_dse + "']]/following-sibling::div//button[@data-toggle='modal'][1]")
        self.update_data_conf_form = (By.XPATH,
                                      "//h4[.//span[.='" + UserData.form_export_name_dse + "']]/following-sibling::div//button[@data-bind='click: emailedExport.updateData']")
        self.update_data_case = (By.XPATH,
                                 "//h4[.//span[.='" + UserData.case_export_name_dse + "']]/following-sibling::div//button[@data-toggle='modal'][1]")
        self.update_data_conf_case = (By.XPATH,
                                      "//h4[.//span[.='" + UserData.case_export_name_dse + "']]/following-sibling::div//button[@data-bind='click: emailedExport.updateData']")

        self.copy_dashfeed_link = (By.XPATH, "(//span[contains(@data-bind, 'copyLinkRequested')])[1]")
        self.dashboard_feed_link = (
            By.XPATH, "//span[@class='input-group-btn']//preceding::a[@class='btn btn-info btn-xs']")
        self.check_data = (By.XPATH, "//*[contains(text(), '@odata.context')]")

        # Power BI / Tableau Integration, Form
        self.powerBI_tab_int = (By.LINK_TEXT, 'PowerBi/Tableau Integration')
        self.copy_odata_link_btn_form = (
        By.XPATH, "//div[./span[text()='" + UserData.odata_feed_form + "']]/following::div[@class='input-group']//a")
        self.copy_odata_link_form = (
        By.XPATH, "//div[./span[text()='" + UserData.odata_feed_form + "']]/following::div[@class='input-group']/input")
        self.copy_odata_link_btn_case = (
        By.XPATH, "//div[./span[text()='" + UserData.odata_feed_case + "']]/following::div[@class='input-group']//a")
        self.copy_odata_link_case = (
        By.XPATH, "//div[./span[text()='" + UserData.odata_feed_case + "']]/following::div[@class='input-group']/input")

        self.edit_button_case = (By.XPATH,
                                 "(//span[contains(text(), 'Copy & Edit Feed')])")
        self.edit_button_form = (By.XPATH,
                                 "(//span[text()='" + UserData.odata_feed_form + "']//following::a[@data-bind='attr: {href: editUrl}'])[1]")
        self.select_none = (By.XPATH, "(//a[@data-bind='click: table.selectNone'])[1]")
        self.first_checkbox = (By.XPATH, "(//input[@type='checkbox'])[3]")
        self.third_checkbox = (By.XPATH, "(//input[@type='checkbox'])[5]")
        self.failed_to_export = (By.XPATH, "//div[@class='alert alert-danger']")

        # bulk export delete
        self.empty_export_block = (By.XPATH, "(//div[@data-bind='visible: showEmpty'])[1]")
        self.select_all_btn = (By.XPATH, '//button[@data-bind="click: selectAll"]')
        self.delete_selected_exports = (By.XPATH, '//a[@href= "#bulk-delete-export-modal"]')
        self.bulk_delete_confirmation_btn = (By.XPATH, '//button[@data-bind="click: BulkExportDelete"]')
        self.alert_button_accept = (By.ID, "hs-eu-confirmation-button")

        # Export Modal
        self.app_type = (By.ID, "id_app_type")
        self.application = (By.ID, "id_application")
        self.module = (By.ID, "id_module")
        self.form = (By.ID, "id_form")
        self.case = (By.ID, "id_case_type")
        self.case_type = (By.XPATH, "//select[@name='case_type']")
        self.model = (By.ID, "id_model_type")

        # Import From Excel
        self.to_be_edited_file = os.path.abspath(
            os.path.join(UserData.USER_INPUT_BASE_DIR, "test_data/import_parent_child_case.xlsx"))

    def get_url_paste_browser(self, username, password, item):
        if item == 'cases':
            odata_feed_link = self.wait_to_get_value(self.copy_odata_link_case)
        elif item == 'forms':
            odata_feed_link = self.wait_to_get_value(self.copy_odata_link_form)
        final_URL_case = f"https://{username}:{password}@{odata_feed_link[8:]}"
        self.driver.get(final_URL_case)

    def date_filter(self):
        self.wait_and_sleep_to_click(self.date_range)
        print(self.date_having_submissions)
        self.wait_to_clear_and_send_keys(self.date_range, self.date_having_submissions)
        self.wait_and_sleep_to_click(self.apply)

    def prepare_and_download_export(self):
        self.wait_and_sleep_to_click(self.export_form_case_data_button)
        self.date_filter()
        self.wait_and_sleep_to_click(self.prepare_export_button)
        try:
            self.wait_till_progress_completes("exports")
            self.wait_for_element(self.download_button, 300)
            self.wait_to_click(self.download_button)
        except TimeoutException:
            if self.is_visible_and_displayed(self.failed_to_export):
                self.driver.refresh()
                self.wait_and_sleep_to_click(self.prepare_export_button)
                self.wait_till_progress_completes("exports")
                self.wait_for_element(self.download_button, 300)
                self.wait_to_click(self.download_button)
        time.sleep(5)
        print("Download form button clicked")

    def find_data_by_id_and_verify(self, row, value, export_name, name_on_hq):
        newest_file = latest_download_file()
        print("Newest file:" + newest_file)
        self.assert_downloaded_file(newest_file, export_name)
        self.wait_to_click(self.find_data_by_ID)
        data = pd.read_excel(newest_file)
        df = pd.DataFrame(data, columns=[row, value])
        ID = df[value].values[0]
        print("ID: ",ID)
        woman_name_excel = df[row].values[0]
        self.wait_to_clear_and_send_keys(self.find_data_by_ID_textbox, str(ID))
        self.wait_and_sleep_to_click(self.find_data_by_ID_button)
        self.wait_for_element(self.view_FormID_CaseID)
        link = self.get_attribute(self.view_FormID_CaseID, "href")
        print(link)
        self.driver.get(link)
        self.is_visible_and_displayed(self.woman_case_name_HQ)
        womanName_HQ = self.wait_to_get_text(name_on_hq)
        assert woman_name_excel == womanName_HQ
        print("Downloaded file has the required data!")

    # Test Case 20_a - Verify Export functionality for Forms

    def add_form_exports(self):
        self.wait_for_element(self.add_export_button, 100)
        self.delete_bulk_exports()
        self.wait_and_sleep_to_click(self.add_export_button)
        time.sleep(30)
        self.wait_for_element(self.app_type, 200)
        self.is_clickable(self.app_type)
        self.select_by_text(self.app_type, UserData.app_type)
        self.select_by_text(self.application, UserData.village_application)
        self.select_by_text(self.module, UserData.case_list_name)
        self.select_by_text(self.form, UserData.form_name)
        self.wait_to_click(self.add_export_conf)
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.form_export_name)
        self.wait_to_click(self.export_settings_create)
        print("Export created!!")

    def form_exports(self):
        self.prepare_and_download_export()
        self.find_data_by_id_and_verify('form.womans_name', 'formid', UserData.form_export_name,
                                        self.woman_form_name_HQ)

    # Test Case 20_b - Verify Export functionality for Cases

    def add_case_exports(self):
        self.wait_to_click(self.export_case_data_link)
        self.wait_for_element(self.add_export_button, 100)
        self.delete_bulk_exports()
        self.wait_and_sleep_to_click(self.add_export_button)
        time.sleep(30)
        self.wait_for_element(self.case_type, 200)
        # self.is_clickable(self.application)
        # self.select_by_text(self.application, UserData.village_application)
        # self.select_by_text(self.case, UserData.case_pregnancy)
        self.select_by_text(self.case, UserData.case_reassign)
        self.wait_to_click(self.add_export_conf)
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.case_export_name)
        self.wait_to_click(self.export_settings_create)
        print("Export created!!")

    def case_exports(self):
        self.wait_and_sleep_to_click(self.export_case_data_link)
        self.prepare_and_download_export()
        self.find_data_by_id_and_verify('name', 'caseid', UserData.case_export_name, self.woman_case_name_HQ)

    # Test Case 21 - Export SMS Messages

    def sms_exports(self):
        self.wait_and_sleep_to_click(self.export_sms_link)
        self.prepare_and_download_export()
        newest_file = latest_download_file()
        print("Newest:", newest_file)
        self.assert_downloaded_file(newest_file, "Messages")
        print("SMS Export successful")

    def create_dse_and_download(self, exported_file, type):
        self.wait_and_sleep_to_click(self.create_DSE_checkbox)
        self.wait_and_sleep_to_click(self.export_settings_create)
        if type == "form":
            self.wait_and_sleep_to_click(self.update_data_form)
            self.wait_and_sleep_to_click(self.update_data_conf_form)
            self.wait_till_progress_completes("integration")
            try:
                assert self.is_present_and_displayed(self.data_upload_msg_form), "Form Export not completed!"
                self.driver.refresh()
                self.wait_to_click(self.download_dse_form)
            except:
                self.driver.refresh()
                self.wait_and_sleep_to_click(self.download_dse_form)
        elif type == "case":
            self.wait_till_progress_completes("integration")
            self.wait_and_sleep_to_click(self.update_data_case)
            self.wait_and_sleep_to_click(self.update_data_conf_case)
            try:
                assert self.is_present_and_displayed(self.data_upload_msg_case), "Case Export not completed!"
                self.driver.refresh()
                self.wait_to_click(self.download_dse_case)
            except:
                self.driver.refresh()
                self.wait_and_sleep_to_click(self.download_dse_case)
        time.sleep(5)
        newest_file = latest_download_file()
        print("Newest:", newest_file)
        print("Exported:", exported_file)
        self.assert_downloaded_file(newest_file, exported_file)

    def cleanup_existing_dse(self):
        # Cleanup existing exports
        self.wait_and_sleep_to_click(self.daily_saved_export_link)
        time.sleep(5)
        self.delete_bulk_exports()

    # Test Case 24_a - Daily saved export, form
    def daily_saved_exports_form(self):
        self.wait_to_click(self.export_form_data_link)
        try:
            self.click(self.edit_form_case_export)
        except (NoSuchElementException, StaleElementReferenceException):
            self.add_form_exports()
            self.wait_and_sleep_to_click(self.edit_form_case_export)
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.form_export_name_dse)
        self.create_dse_and_download(UserData.form_export_name_dse, "form")
        print("DSE Form Export successful")

    # Test Case 24_b - Daily saved export, case
    def daily_saved_exports_case(self):
        self.wait_to_click(self.export_case_data_link)
        try:
            self.click(self.edit_form_case_export)
        except NoSuchElementException:
            self.add_case_exports()
            self.wait_and_sleep_to_click(self.edit_form_case_export)
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.case_export_name_dse)
        self.create_dse_and_download(UserData.case_export_name_dse, "case")
        print("DSE Case Export successful")

    # Test Case - 25 - Excel Dashboard Integration, form
    def excel_dashboard_integration_form(self):
        self.wait_and_sleep_to_click(self.export_excel_dash_int)
        self.wait_and_sleep_to_click(self.add_export_button)
        time.sleep(10)
        self.is_visible_and_displayed(self.model)
        self.wait_for_element(self.model, 100)
        self.select_by_value(self.model, UserData.model_type_form)
        self.select_by_text(self.app_type, UserData.app_type)
        self.select_by_text(self.application, UserData.village_application)
        self.select_by_text(self.module, UserData.case_list_name)
        self.select_by_text(self.form, UserData.form_name)
        self.wait_and_sleep_to_click(self.add_export_conf)
        print("Dashboard Feed added!!")
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.dashboard_feed_form)
        self.click(self.export_settings_create)
        print("Dashboard Form Feed created!!")
        self.wait_and_sleep_to_click(self.update_data)
        self.wait_till_progress_completes("integration")
        self.wait_and_sleep_to_click(self.update_data_conf)
        assert self.is_visible_and_displayed(self.data_upload_msg), "Export not completed!"
        self.driver.refresh()

    # Test Case - 26 - Excel Dashboard Integration, case

    def excel_dashboard_integration_case(self):
        self.wait_and_sleep_to_click(self.export_excel_dash_int)
        self.wait_and_sleep_to_click(self.add_export_button)
        time.sleep(50)
        self.is_visible_and_displayed(self.model, 200)
        self.wait_for_element(self.model, 100)
        self.select_by_value(self.model, UserData.model_type_case)
        try:
            self.select_by_text(self.application, UserData.village_application)
        except:
            print("Application dropdown is not present")
        self.select_by_text(self.case, UserData.case_pregnancy)
        self.wait_and_sleep_to_click(self.add_export_conf)
        print("Dashboard Feed added!!")
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.dashboard_feed_case)
        self.click(self.export_settings_create)
        print("Dashboard Form Feed created!!")
        self.wait_and_sleep_to_click(self.update_data)
        self.wait_till_progress_completes("integration")
        self.wait_and_sleep_to_click(self.update_data_conf)
        assert self.is_visible_and_displayed(self.data_upload_msg), "Export not completed!"
        self.driver.refresh()

    def check_feed_link(self):
        try:
            self.wait_and_sleep_to_click(self.copy_dashfeed_link)
            dashboard_feed_link = self.get_attribute(self.dashboard_feed_link, "href")
            print(dashboard_feed_link)
            # self.switch_to_new_tab()
            self.driver.get(dashboard_feed_link)
            dashboard_feed_data = self.driver.page_source
            if dashboard_feed_data != "":
                print("Excel Dashboard has data")
            else:
                print("Excel Dashboard is empty")
            # self.driver.close()
            self.driver.back()
            return dashboard_feed_link
        except StaleElementReferenceException:
            print(StaleElementReferenceException)

    # Test Case - 28 - Power BI / Tableau Integration, Form

    def power_bi_tableau_integration_form(self, username, password):
        try:
            self.wait_and_sleep_to_click(self.powerBI_tab_int)
        except ElementClickInterceptedException:
            self.js_click(self.powerBI_tab_int)
        self.delete_bulk_exports()
        self.wait_and_sleep_to_click(self.add_export_button)
        time.sleep(30)
        self.is_visible_and_displayed(self.model)
        self.wait_for_element(self.model, 100)
        self.select_by_value(self.model, UserData.model_type_form)
        self.select_by_text(self.app_type, UserData.app_type)
        self.select_by_text(self.application, UserData.village_application)
        self.select_by_text(self.module, UserData.case_list_name)
        self.select_by_text(self.form, UserData.form_name)
        self.wait_and_sleep_to_click(self.add_export_conf)
        print("Odata form Feed added!!")
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.odata_feed_form)
        self.click(self.export_settings_create)
        print("Odata Form Feed created!!")
        self.driver.refresh()
        self.wait_and_sleep_to_click(self.copy_odata_link_btn_form)
        self.get_url_paste_browser(username, password, "forms")
        self.assert_odata_feed_data()

    # Test Case - 27 - Power BI / Tableau Integration, Case`
    def power_bi_tableau_integration_case(self, username, password):
        self.driver.refresh()
        self.wait_to_click(self.powerBI_tab_int)
        self.delete_bulk_exports()
        self.wait_and_sleep_to_click(self.add_export_button)
        time.sleep(30)
        self.is_visible_and_displayed(self.model, 200)
        self.wait_for_element(self.model, 100)
        self.select_by_value(self.model, UserData.model_type_case)
        try:
            self.select_by_text(self.application, UserData.village_application)
        except:
            print("Application dropdown is not present")
        self.select_by_text(self.case, UserData.case_pregnancy)
        self.wait_and_sleep_to_click(self.add_export_conf)
        print("Odata Case Feed added!!")
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.odata_feed_case)
        # selcting first three property
        self.wait_and_sleep_to_click(self.select_none)
        self.wait_and_sleep_to_click(self.first_checkbox)
        self.wait_and_sleep_to_click(self.third_checkbox)
        # saving export
        self.click(self.export_settings_create)
        print("Odata Case Feed created!!")
        self.driver.refresh()
        self.wait_and_sleep_to_click(self.copy_odata_link_btn_case)
        self.get_url_paste_browser(username, password, "cases")
        self.assert_odata_feed_data()

    def delete_bulk_exports(self):
        try:
            self.wait_to_click(self.select_all_btn)
            self.wait_to_click(self.delete_selected_exports)
            self.wait_to_click(self.bulk_delete_confirmation_btn)
        except TimeoutException:
            print("No exports available")

    def delete_all_bulk_exports(self):
        self.wait_and_sleep_to_click(self.export_form_data_link)
        self.delete_bulk_exports()
        print("Bulk exports deleted for Export Form data")
        self.wait_and_sleep_to_click(self.export_case_data_link)
        self.delete_bulk_exports()
        print("Bulk exports deleted for Export Case data")
        self.wait_and_sleep_to_click(self.daily_saved_export_link)
        self.delete_bulk_exports()
        print("Bulk exports deleted for Daily Saved Export Cases")

    def delete_all_bulk_integration_exports(self):
        self.wait_and_sleep_to_click(self.powerBI_tab_int)
        self.delete_bulk_exports()
        print("Bulk exports deleted for Power BI Reports")
        self.wait_and_sleep_to_click(self.export_excel_dash_int)
        self.delete_bulk_exports()
        print("Bulk exports deleted for Export Excel Int Reports")

    def assert_odata_feed_data(self):
        odata_feed_data = self.driver.page_source
        verify_data = self.find_elements(self.check_data)
        assert len(verify_data) > 0, "Odata feed is Empty "
        # self.driver.close()  # Close the feed URL
        self.driver.back()

    def add_updated_case_exports(self):
        print("Sleeping for some time for the cases to be updated in the exports")
        time.sleep(200)
        self.wait_to_click(self.export_case_data_link)
        self.delete_bulk_exports()
        time.sleep(5)
        self.wait_and_sleep_to_click(self.add_export_button)
        self.wait_for_element(self.case_type, 200)
        self.select_by_text(self.case, UserData.case_update_name)
        self.wait_to_click(self.add_export_conf)
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.case_updated_export_name)
        self.wait_to_click(self.export_settings_create)
        print("Export created!!")
        self.wait_to_click(self.export_button)
        time.sleep(3)
        self.wait_for_element(self.prepare_export_button)
        if self.is_present(self.web_users_option):
            print("Web Users is already selected")
        else:
            self.send_keys(self.users_field, UserData.web_user)
            self.wait_to_click((By.XPATH, self.user_from_list.format(UserData.web_user)))
            print("Selecting Web Users")
            time.sleep(2)
            ActionChains(self.driver).send_keys(Keys.TAB).perform()
        if self.is_present(self.all_data_option):
            print("Web Users is already selected")
        else:
            self.send_keys(self.users_field, UserData.all_data)
            self.wait_to_click((By.XPATH, self.user_from_list.format(UserData.all_data)))
            print("Selecting Web Users")
            time.sleep(2)
            ActionChains(self.driver).send_keys(Keys.TAB).perform()
        self.wait_to_clear_and_send_keys(self.date_range, self.next_date_range + Keys.TAB)
        self.wait_and_sleep_to_click(self.prepare_export_button)
        time.sleep(10)
        try:
            self.wait_till_progress_completes("exports")
            self.wait_for_element(self.download_button, 300)
            self.wait_to_click(self.download_button)
            time.sleep(5)
        except TimeoutException:
            if self.is_visible_and_displayed(self.failed_to_export):
                self.driver.refresh()
                self.wait_and_sleep_to_click(self.prepare_export_button)
                self.wait_till_progress_completes("exports")
                self.wait_for_element(self.download_button, 300)
                self.wait_to_click(self.download_button)
                time.sleep(5)
        print("Download form button clicked")

    def verify_export_has_updated_case_data(self, case_id, case_name, value):
        print(case_id, case_name, value)
        newest_file = latest_download_file()
        print("Newest file:" + newest_file)
        self.assert_downloaded_file(newest_file, UserData.case_updated_export_name)
        data = pd.read_excel(newest_file)
        df = pd.DataFrame(data, columns=[UserData.case_id, UserData.text_value, UserData.random_value])
        case_id_row = df[df[UserData.case_id] == case_id].index[0]
        name_in_file = df[UserData.text_value].loc[case_id_row]
        value_in_file = df[UserData.random_value].loc[case_id_row]
        print(case_id_row, name_in_file, value_in_file)
        assert str(value_in_file) == value and str(name_in_file) == case_name
        print("Downloaded file has the required data!")

    def clean_up_case_data(self):
        self.wait_and_sleep_to_click(self.export_case_data_link)
        self.delete_bulk_exports()
        print("Bulk exports deleted for Export Case data")

    def verify_duplicate_data_in_dashboard(self, link, username, password):
        print(link)
        resp = requests.get(link, auth=(username, password)).text
        data = pd.read_html(resp, flavor='html5lib')
        data = (pd.DataFrame(data[0])).reset_index()
        duplicate = data[data.duplicated()]
        if len(duplicate)>0:
            print(duplicate)
        else:
            print("No duplicate data present")


    def add_case_exports(self):
        self.wait_to_click(self.export_case_data_link)
        self.delete_bulk_exports()
        self.wait_and_sleep_to_click(self.add_export_button)
        self.wait_for_element(self.case_type, 200)
        self.select_by_text(self.case, UserData.case_reassign)
        self.wait_to_click(self.add_export_conf)
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.case_export_name)
        self.wait_to_click(self.export_settings_create)
        print("Export created!!")


    def add_form_exports_reassign(self):
        self.delete_bulk_exports()
        self.wait_and_sleep_to_click(self.add_export_button)
        self.wait_for_element(self.app_type, 200)
        self.select_by_text(self.app_type, UserData.app_type)
        self.select_by_text(self.application, UserData.reassign_cases_application)
        self.select_by_text(self.module, UserData.case_list_name)
        self.select_by_text(self.form, UserData.form_name)
        self.wait_to_click(self.add_export_conf)
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.p1p2_form_export_name)
        self.wait_to_click(self.export_settings_create)
        print("Export created!!")
        self.download_export_without_condition("form")
        newest_file = latest_download_file()
        print("Newest file:" + newest_file)
        self.assert_downloaded_file(newest_file, UserData.p1p2_form_export_name)
        return newest_file

    def verify_export_count(self, name):
        data = pd.read_excel(name)
        df = pd.DataFrame(data)
        rows_count = df.shape[0]
        print(int(rows_count))
        assert int(rows_count) >= 2000, "Export is not showing all the data"
        print("Export is successfully loading more than 2000 rows of data")

    def download_export_without_condition(self, type):
        self.wait_and_sleep_to_click(self.export_form_case_data_button)
        self.wait_for_element(self.prepare_export_button)
        if type == "form":
            if self.is_present(self.web_users_option):
                print("Web Users is already selected")
            else:
                self.send_keys(self.users_field, UserData.web_user)
                self.wait_to_click((By.XPATH, self.user_from_list.format(UserData.web_user)))
                print("Selecting Web Users")
                time.sleep(2)
                ActionChains(self.driver).send_keys(Keys.TAB).perform()
        else:
            if self.is_present(self.all_data_option):
                print("Web Users is already selected")
            else:
                self.send_keys(self.users_field, UserData.all_data)
                self.wait_to_click((By.XPATH, self.user_from_list.format(UserData.all_data)))
                print("Selecting Web Users")
                time.sleep(2)
                ActionChains(self.driver).send_keys(Keys.TAB).perform()
        self.wait_and_sleep_to_click(self.prepare_export_button)
        try:
            self.wait_till_progress_completes("exports")
            self.wait_for_element(self.download_button, 300)
            self.wait_to_click(self.download_button)
        except TimeoutException:
            if self.is_visible_and_displayed(self.failed_to_export):
                self.driver.refresh()
                self.wait_and_sleep_to_click(self.prepare_export_button)
                self.wait_till_progress_completes("exports")
                self.wait_for_element(self.download_button, 300)
                self.wait_to_click(self.download_button)
        time.sleep(5)
        print("Download form button clicked")

    def add_case_exports_reassign(self):
        self.wait_to_click(self.export_case_data_link)
        self.delete_bulk_exports()
        self.wait_and_sleep_to_click(self.add_export_button)
        self.wait_for_element(self.case_type, 200)
        self.select_by_text(self.case, UserData.case_reassign)
        self.wait_to_click(self.add_export_conf)
        self.wait_for_element(self.export_name)
        self.wait_to_clear_and_send_keys(self.export_name, UserData.p1p2_case_export_name)
        self.wait_to_click(self.export_settings_create)
        print("Export created!!")
        self.download_export_without_condition("case")
        newest_file = latest_download_file()
        print("Newest file:" + newest_file)
        self.assert_downloaded_file(newest_file, UserData.p1p2_case_export_name)
        return newest_file


    def check_for_related_cases(self, parent_id):
        self.wait_for_element(self.find_data_by_ID)
        self.wait_to_click(self.find_data_by_ID)
        self.wait_to_clear_and_send_keys(self.find_data_by_case_ID_textbox, parent_id)
        self.wait_and_sleep_to_click(self.find_data_by_case_ID_button)
        self.wait_for_element(self.view_FormID_CaseID)
        link = self.get_attribute(self.view_FormID_CaseID, "href")
        print(link)
        self.driver.get(link)
        self.wait_for_element((By.XPATH, self.case_id_value.format(parent_id)))
        if self.is_present(self.related_cases_tab):
            self.validate_child_case_data()
            return "assign to parent 2"
        else:
            return "assign to parent 1"


    def validate_child_case_data(self):
        self.wait_to_click(self.related_cases_tab)
        self.wait_for_element(self.related_cases_view)
        self.wait_to_click(self.related_cases_view)
        time.sleep(5)
        self.wait_for_element((By.XPATH, self.case_id_value.format(UserData.child_case_id)))

    def prepare_parent_child_import_excel(self, text):
        if text == "assign to parent 2":
            print("Preparing excel for assignment to Parent: ", UserData.parent_2_id)
            workbook = load_workbook(filename=self.to_be_edited_file)
            sheet = workbook.active
            sheet["B2"] = UserData.parent_2_id
            sheet.title = "Sheet 1"
            filename = os.path.abspath(
            os.path.join(UserData.USER_INPUT_BASE_DIR, "test_data/import_to_parent_" + UserData.parent_2_id+".xlsx"))
            workbook.save(filename=filename)
            print(filename)
            return filename
        else:
            print("Preparing excel for assignment to Parent: ", UserData.parent_1_id)
            workbook = load_workbook(filename=self.to_be_edited_file)
            sheet = workbook.active
            sheet["B2"] = UserData.parent_1_id
            sheet.title = "Sheet 1"
            filename = os.path.abspath(
                os.path.join(UserData.USER_INPUT_BASE_DIR,
                             "test_data/import_to_parent_" + UserData.parent_1_id + ".xlsx"))
            workbook.save(filename=filename)
            return filename

    def verify_case_import(self, text):
        self.wait_for_element(self.find_data_by_ID)
        self.js_click(self.find_data_by_ID)
        if text == "assign to parent 2":
            parent_id = UserData.parent_2_id
        else:
            parent_id = UserData.parent_1_id
        self.wait_for_element(self.find_data_by_case_ID_textbox)
        self.wait_to_clear_and_send_keys(self.find_data_by_case_ID_textbox, parent_id)
        self.wait_to_click(self.find_data_by_case_ID_button)
        self.wait_for_element(self.view_FormID_CaseID)
        link = self.get_attribute(self.view_FormID_CaseID, "href")
        print(link)
        self.driver.get(link)
        self.wait_for_element((By.XPATH, self.case_id_value.format(parent_id)))
        assert self.is_present(self.related_cases_tab), "Parent not reassigned"
        self.validate_child_case_data()

